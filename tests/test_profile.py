import os
import pathlib
import openpyxl
import pytest
from pages.profile_page import AccountPage

excel_path = os.path.join(pathlib.Path(__file__).resolve().parents[1], "data", "Timkiemdata.xlsx")
sheet_name = "Ttcn"

def read_account_data(path=excel_path, sheet=sheet_name):
    wb = openpyxl.load_workbook(path)
    if sheet not in wb.sheetnames:
        raise RuntimeError(f"Sheet '{sheet}' không tồn tại trong {path}")
    ws = wb[sheet]
    headers = [c.value.strip() for c in ws[1]]
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        item = {}
        for idx, val in enumerate(row):
            key = headers[idx] if idx < len(headers) else f"col{idx+1}"
            item[key.strip()] = "" if val is None else str(val).strip()
        data.append(item)
    wb.close()
    return data

DATA = read_account_data()

def pick(row, possible_names):
    for name in possible_names:
        if name in row and row[name] not in (None, ""):
            return row[name]
        for k in row:
            if k.lower() == name.lower() and row[k] not in (None, ""):
                return row[k]
    return ""

@pytest.mark.parametrize("row", DATA)
def test_account_info_flow(driver, row):
    phone = pick(row, ["Phone"])
    password = pick(row, ["Password"])
    fullname = pick(row, ["FullName"])
    email = pick(row, ["Email"])
    address = pick(row, ["Address"])
    expected_title = pick(row, ["Expected"])

    p = AccountPage(driver)
    p.open_home()
    p.login(phone, password)
    p.open_account_page()
    p.update_account(fullname, email, address)

    actual_title = p.get_popup_title()
    print(f"Expected: {expected_title}")
    print(f"Actual  : {actual_title}")
    assert  expected_title in actual_title
