# utils/excel_logger.py
import openpyxl
from datetime import datetime
import os

class ExcelLogger:
    def __init__(self, file_path="logs/test_results.xlsx", sheet_name="LoginTest"):
        self.file_path = file_path
        self.sheet_name = sheet_name

        # Tạo thư mục nếu chưa có
        os.makedirs(os.path.dirname(self.file_path), exist_ok=True)

        # Tạo file Excel nếu chưa tồn tại
        if not os.path.exists(self.file_path):
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = self.sheet_name
            sheet.append(["Thời gian", "Test Case", "Kết quả", "Ghi chú"])
            workbook.save(self.file_path)

    def log_result(self, test_case, result, note=""):
        workbook = openpyxl.load_workbook(self.file_path)

        # Nếu sheet chưa có, tạo mới
        if self.sheet_name not in workbook.sheetnames:
            workbook.create_sheet(self.sheet_name)
        sheet = workbook[self.sheet_name]

        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        sheet.append([timestamp, test_case, result, note])
        workbook.save(self.file_path)
